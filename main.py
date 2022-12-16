from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill  # Módulo de llenado de importación
nameFile=input("Ingrese el nombre del archivo con su extencion: ")
wb = load_workbook(filename = nameFile)
ft = Font(color="FF0000")
nameSheet=input("Ingrese el nombre de la hoja del libro: ")
sheet_ranges = wb[nameSheet]
columnData=input("Ingrese la letra de la columna a buscar repetidos en MAYUSCULA: ")
columna= (columnData)
numColumna= 2
evaColumna= (columnData)
evanumColumna= 3
evanumColumna1= 3
numberColumnData=int(input("Ingrese el numero de la columna a resaltar: "))

Color = ['00FFFF00', '000000']  # Amarillo y negro

for i in sheet_ranges:
    value=(sheet_ranges[columna+str(numColumna)].value)
    if value==None:
        break
    for j in sheet_ranges:
        value1=(sheet_ranges[evaColumna+str(evanumColumna)].value)
        evanumColumna += 1
        fille = PatternFill('solid', fgColor=Color[0])  # Establecer el color de relleno en naranja
        if value1==value:
            print (value)
            sheet_ranges.cell(row=evanumColumna-1, column=numberColumnData, ).fill = fille
        if value1==None:
            evanumColumna= numColumna+2
            break
    numColumna += 1

wb.save("prueba.xlsx")        
print("Finished")
